# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║   ROBÔ DE CONSULTA PROCESSUAL — SEBRAE CONTENCIOSO  (versão janela/GUI)     ║
║   API DataJud CNJ — Elasticsearch por tribunal                             ║
║                                                                              ║
║   • Janela simples: escolher o CSV, clicar em Iniciar, ver o progresso     ║
║   • A CHAVE da API é buscada automaticamente no site do CNJ a cada uso     ║
║     (assim o robô continua funcionando mesmo quando o CNJ troca a chave)   ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import re
import json
import time
import unicodedata
import threading
import webbrowser
from datetime import datetime
from pathlib import Path

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import requests
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configurações ──────────────────────────────────────────────────────────
API_BASE = "https://api-publica.datajud.cnj.jus.br"

# Página oficial do CNJ de onde a chave pública é lida automaticamente:
URL_CHAVE_CNJ = "https://datajud-wiki.cnj.jus.br/api-publica/acesso/"

# Última chave pública conhecida (usada só se a leitura automática falhar).
# Atualizada em 17/07/2026.
CHAVE_FALLBACK = "cDZHYzlZa0JadVREZDJCendQbXY6SkJlTzNjLV9TRENyQk1RdnFKZGRQdw=="

TIMEOUT = 15
PAUSA = 0.65      # ~92 req/min, abaixo do limite de 100
PAUSA_429 = 10.0
MAX_TENT = 3

# Códigos de movimento do CNJ (TPU) que indicam ARQUIVAMENTO DEFINITIVO.
#   246 = "Definitivo"  (Arquivamento Definitivo)
#    22 = "Baixa Definitiva"
# Confirmados empiricamente na base do DataJud com os processos desta lista.
CODIGOS_ARQUIVAMENTO_DEFINITIVO = {22, 246}

# Mapeamento de número CNJ → índice DataJud
INDICE_MAP = {
    "4.01": "trf1", "4.02": "trf2", "4.03": "trf3", "4.04": "trf4",
    "4.05": "trf5", "4.06": "trf6",
    "5.01": "trt1", "5.02": "trt2", "5.03": "trt3", "5.04": "trt4",
    "5.05": "trt5", "5.06": "trt6", "5.07": "trt7", "5.08": "trt8",
    "5.09": "trt9", "5.10": "trt10", "5.11": "trt11", "5.12": "trt12",
    "5.13": "trt13", "5.14": "trt14", "5.15": "trt15", "5.17": "trt17",
    "5.18": "trt18", "5.20": "trt20", "5.21": "trt21", "5.22": "trt22",
    "5.23": "trt23", "5.24": "trt24",
    "8.01": "tjac", "8.02": "tjal", "8.03": "tjap", "8.04": "tjam",
    "8.05": "tjba", "8.06": "tjce", "8.07": "tjdft", "8.08": "tjes",
    "8.09": "tjgo", "8.10": "tjma", "8.11": "tjmt", "8.12": "tjms",
    "8.13": "tjmg", "8.14": "tjpa", "8.15": "tjpb", "8.16": "tjpr",
    "8.17": "tjpe", "8.18": "tjpi", "8.19": "tjrj", "8.20": "tjrn",
    "8.21": "tjrs", "8.22": "tjro", "8.23": "tjrr", "8.24": "tjsc",
    "8.25": "tjsp", "8.26": "tjsp", "8.27": "tjse", "8.28": "tjto",
    "1.00": "stf", "2.00": "stj", "3.00": "tst",
}


# ── Chave de API (auto-atualização) ─────────────────────────────────────────
def obter_chave_api(logfn=print) -> str:
    """
    Busca a chave pública ATUAL no site do CNJ.
    Se não conseguir (sem internet, site fora do ar, mudança de layout),
    usa a última chave conhecida (CHAVE_FALLBACK).
    """
    try:
        r = requests.get(URL_CHAVE_CNJ, timeout=TIMEOUT)
        if r.status_code == 200:
            # A chave é um token base64 longo (termina com '=' ou '==').
            candidatos = re.findall(r"[A-Za-z0-9+/]{50,}={0,2}", r.text)
            for cand in candidatos:
                if cand.endswith("="):  # chaves do CNJ terminam com padding '='
                    logfn("Chave pública obtida automaticamente do site do CNJ.")
                    return cand
        logfn("Não foi possível ler a chave no site do CNJ; usando a última conhecida.")
    except Exception as e:
        logfn(f"Sem acesso ao site do CNJ ({type(e).__name__}); usando a última chave conhecida.")
    return CHAVE_FALLBACK


def validar_chave(chave: str, logfn=print) -> bool:
    """Faz uma chamada de teste para confirmar que a chave está ativa."""
    headers = {"Authorization": f"APIKey {chave}", "Content-Type": "application/json"}
    try:
        r = requests.post(f"{API_BASE}/api_publica_tjsp/_search",
                          headers=headers, json={"size": 1, "query": {"match_all": {}}},
                          timeout=TIMEOUT)
        if r.status_code == 200:
            logfn("Chave da API validada com sucesso (ativa).")
            return True
        logfn(f"Chave rejeitada pelo servidor (HTTP {r.status_code}).")
        return False
    except Exception as e:
        logfn(f"Não foi possível validar a chave agora ({type(e).__name__}).")
        return False


# ── Lógica de consulta ──────────────────────────────────────────────────────
def indice_datajud(cnj: str):
    m = re.match(r"\d{7}-\d{2}\.\d{4}\.(\d)\.(\d{2})\.", str(cnj))
    if not m:
        return None
    nome = INDICE_MAP.get(f"{m.group(1)}.{m.group(2)}")
    return f"api_publica_{nome}" if nome else None


def _sem_acento(s) -> str:
    s = str(s or "").lower()
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def _ultimo_movimento(movs):
    """Retorna o movimento mais recente (por data/hora)."""
    if not movs:
        return None
    return sorted(movs, key=lambda m: str(m.get("dataHora", "")))[-1]


def is_arquivado(dados: dict):
    """
    Detecta arquivamento definitivo pelo ÚLTIMO movimento do processo.
    Prioriza o código do movimento (CNJ/TPU); usa o nome como reforço.
    """
    ult = _ultimo_movimento(dados.get("movimentos", []) or [])
    if not ult:
        return False, "Processo sem movimentos registrados"

    cod = ult.get("codigo")
    try:
        cod_int = int(cod)
    except (TypeError, ValueError):
        cod_int = None
    nome = ult.get("nome", "")
    nome_n = _sem_acento(nome)

    if cod_int in CODIGOS_ARQUIVAMENTO_DEFINITIVO:
        return True, f"Último movimento: «{nome}» (cód. {cod})"
    if "baixa definitiva" in nome_n or nome_n == "definitivo" or \
       ("arquiv" in nome_n and "defin" in nome_n):
        return True, f"Último movimento: «{nome}»"
    return False, f"Último movimento: «{nome}» — sem arquivamento definitivo"


def extrair_info(dados: dict):
    movs = dados.get("movimentos", []) or []
    ult = _ultimo_movimento(movs)
    ult_data, ult_mov = "—", "—"
    if ult:
        d = ult.get("dataHora", "")
        ult_data = d[:10] if isinstance(d, str) and len(d) >= 10 else "—"
        ult_mov = str(ult.get("nome", "—"))[:80]
    return {
        "Tribunal": str(dados.get("tribunal", "—"))[:50],
        "Classe": str((dados.get("classe") or {}).get("nome", "—"))[:40],
        "Grau": dados.get("grau", "—"),
        "Data Última Mov.": ult_data,
        "Última Movimentação": ult_mov,
        "Qtd Movimentos": len(movs),
    }


def consultar(cnj, indice, headers, tentativa=1):
    url = f"{API_BASE}/{indice}/_search"
    # No DataJud o numeroProcesso é gravado SÓ com dígitos (sem "-", "." etc.).
    # Consultar com a pontuação retorna zero resultados — por isso normalizamos.
    cnj_digitos = re.sub(r"\D", "", str(cnj))
    body = {
        "size": 1,
        "query": {"term": {"numeroProcesso.keyword": cnj_digitos}},
        "_source": ["numeroProcesso", "classe", "grau", "tribunal",
                    "orgaoJulgador", "movimentos"],
    }
    try:
        r = requests.post(url, headers=headers, json=body, timeout=TIMEOUT)
        if r.status_code == 200:
            hits = r.json().get("hits", {}).get("hits", [])
            if not hits:
                return "NÃO ENCONTRADO", "Processo não localizado neste tribunal no DataJud", {}
            dados = hits[0]["_source"]
            arquivado, motivo = is_arquivado(dados)
            info = extrair_info(dados)
            return ("ARQUIVADO" if arquivado else "EM ANDAMENTO"), motivo, info
        elif r.status_code == 404:
            return "NÃO ENCONTRADO", "Índice/tribunal não disponível no DataJud", {}
        elif r.status_code == 429:
            time.sleep(PAUSA_429)
            if tentativa < MAX_TENT:
                return consultar(cnj, indice, headers, tentativa + 1)
            return "ERRO 429", "Rate limit excedido após retentativas", {}
        elif r.status_code in (401, 403):
            return "ACESSO NEGADO", f"HTTP {r.status_code} — chave inválida ou tribunal restrito", {}
        else:
            if tentativa < MAX_TENT:
                time.sleep(2)
                return consultar(cnj, indice, headers, tentativa + 1)
            return f"ERRO HTTP {r.status_code}", r.text[:100], {}
    except requests.ConnectionError:
        return "ERRO CONEXÃO", "Sem acesso à internet ou API indisponível", {}
    except requests.Timeout:
        if tentativa < MAX_TENT:
            return consultar(cnj, indice, headers, tentativa + 1)
        return "TIMEOUT", "Servidor não respondeu em 15s", {}
    except Exception as e:
        return "ERRO", str(e)[:100], {}


# ── Motor: percorre o CSV e gera o Excel ────────────────────────────────────
def executar_consulta(caminho_csv, headers, progresso_cb, log_cb, parar_flag):
    """
    progresso_cb(feitos, total, contadores) — chamado a cada processo.
    log_cb(texto) — mensagens de log.
    parar_flag() — retorna True se o usuário pediu para cancelar.
    Retorna o caminho do arquivo Excel gerado (ou None se cancelado/erro).
    """
    inicio = datetime.now()
    csv_path = Path(caminho_csv)
    log_cb(f"Lendo arquivo: {csv_path.name}")

    try:
        df = pd.read_csv(csv_path, encoding="utf-8-sig")
    except Exception:
        df = pd.read_csv(csv_path, encoding="latin-1", sep=None, engine="python")
    log_cb(f"{len(df)} linhas carregadas.")

    # Garante a coluna de índice DataJud
    if "Indice_DataJud" not in df.columns or df["Indice_DataJud"].isna().all():
        df["Indice_DataJud"] = df.get("CNJ_Extraido", pd.Series([None] * len(df))).apply(indice_datajud)

    com_cnj, sem_cnj = [], []
    for _, row in df.iterrows():
        cnj = str(row.get("CNJ_Extraido", "") or "").strip()
        idx = str(row.get("Indice_DataJud", "") or "").strip()
        if not cnj or not idx or idx.lower() in ("nan", "none"):
            sem_cnj.append(row)
        else:
            com_cnj.append(row)
    com_cnj = pd.DataFrame(com_cnj)
    sem_cnj = pd.DataFrame(sem_cnj)
    log_cb(f"Consultáveis pela API: {len(com_cnj)} | Sem CNJ padrão: {len(sem_cnj)}")

    saida = []
    contadores = {"ARQUIVADO": 0, "EM ANDAMENTO": 0, "NÃO ENCONTRADO": 0,
                  "ACESSO NEGADO": 0, "ERRO": 0}
    total = len(com_cnj)

    for i, (_, row) in enumerate(com_cnj.iterrows()):
        if parar_flag():
            log_cb("Consulta cancelada pelo usuário.")
            break
        cnj = str(row["CNJ_Extraido"]).strip()
        indice = str(row["Indice_DataJud"]).strip()
        status, descricao, info = consultar(cnj, indice, headers)
        chave = status if status in contadores else "ERRO"
        contadores[chave] = contadores.get(chave, 0) + 1

        saida.append({
            "#": i + 1,
            "Controle Interno": str(row.get("Controle Interno do Sebrae", "")),
            "Estado": str(row.get("Estado", "")),
            "CNJ": cnj,
            "Número Original": str(row.get("Número do Processo", "")),
            "Classe de Ação": str(row.get("Classe de Ação", "")),
            "Índice DataJud": indice,
            "Status": status,
            "Detalhamento": descricao[:120],
            "Tribunal": info.get("Tribunal", "—"),
            "Classe Processo": info.get("Classe", "—"),
            "Grau": info.get("Grau", "—"),
            "Data Última Movimentação": info.get("Data Última Mov.", "—"),
            "Última Movimentação": info.get("Última Movimentação", "—"),
            "Qtd. Movimentos": info.get("Qtd Movimentos", "—"),
            "ARQUIVADO DEFINITIVAMENTE": "✅ SIM" if status == "ARQUIVADO" else (
                "❌ NÃO" if status == "EM ANDAMENTO" else "❓ VERIFICAR"),
        })
        progresso_cb(i + 1, total, contadores)
        time.sleep(PAUSA)

    # Processos sem CNJ padrão
    for j, (_, row) in enumerate(sem_cnj.iterrows()):
        saida.append({
            "#": len(com_cnj) + j + 1,
            "Controle Interno": str(row.get("Controle Interno do Sebrae", "")),
            "Estado": str(row.get("Estado", "")),
            "CNJ": "—",
            "Número Original": str(row.get("Número do Processo", "")),
            "Classe de Ação": str(row.get("Classe de Ação", "")),
            "Índice DataJud": "—",
            "Status": "SEM CNJ PADRÃO",
            "Detalhamento": "Número em formato antigo — consultar portal do tribunal",
            "Tribunal": "—", "Classe Processo": "—", "Grau": "—",
            "Data Última Movimentação": "—", "Última Movimentação": "—",
            "Qtd. Movimentos": "—",
            "ARQUIVADO DEFINITIVAMENTE": "❓ VERIFICAR MANUAL",
        })

    df_res = pd.DataFrame(saida)
    if df_res.empty:
        log_cb("Nenhum resultado gerado.")
        return None

    arquivados = df_res[df_res["ARQUIVADO DEFINITIVAMENTE"] == "✅ SIM"]
    em_andamento = df_res[df_res["ARQUIVADO DEFINITIVAMENTE"] == "❌ NÃO"]
    verificar = df_res[~df_res["ARQUIVADO DEFINITIVAMENTE"].isin(["✅ SIM", "❌ NÃO"])]

    fim = datetime.now()
    dur = int((fim - inicio).total_seconds())
    log_cb(f"Concluído em {dur // 60}min {dur % 60}s — "
           f"Arquivados: {len(arquivados)} | Andamento: {len(em_andamento)} | "
           f"Verificar: {len(verificar)}")

    # Salva ao lado do CSV, com data/hora no nome
    saida_path = csv_path.parent / f"Resultado_Consulta_{fim.strftime('%Y-%m-%d_%H%M')}.xlsx"
    salvar_xlsx(df_res, arquivados, em_andamento, verificar, fim, saida_path)
    log_cb(f"Arquivo gerado: {saida_path.name}")
    return str(saida_path)


# ── SALVAR XLSX ─────────────────────────────────────────────────────────────
def salvar_xlsx(df_res, arquivados, em_andamento, verificar, fim, saida_path):
    NAVY = "1A2B5C"; BLUE = "2E5AAC"; RED = "C0392B"; GREEN = "1A7A4A"
    AMBER = "D97706"; WHITE = "FFFFFF"; GRAY_L = "F4F6FB"; GOLD = "D4A017"

    def fl(c): return PatternFill("solid", fgColor=c)
    def ft(bold=False, color="1A2B5C", size=9):
        return Font(name="Arial", bold=bold, color=color, size=size)
    def al(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def bd():
        s = Side(border_style="thin", color="C5CDE8")
        return Border(left=s, right=s, top=s, bottom=s)

    wb = openpyxl.Workbook()
    agora = fim.strftime("%d/%m/%Y %H:%M")
    colunas = df_res.columns.tolist()
    larguras = [4, 18, 6, 26, 34, 22, 22, 18, 50, 30, 28, 6, 18, 40, 14, 22]
    larguras += [15] * (len(colunas) - len(larguras))

    def escrever_aba(ws, titulo, df_tab, cor):
        ws.sheet_view.showGridLines = False
        ws.merge_cells(f"A1:{get_column_letter(len(colunas))}1")
        c = ws["A1"]; c.value = titulo
        c.font = Font(name="Arial", bold=True, color=WHITE, size=12)
        c.fill = fl(cor); c.alignment = al("center"); ws.row_dimensions[1].height = 28
        ws.merge_cells(f"A2:{get_column_letter(len(colunas))}2")
        c = ws["A2"]
        c.value = f"{len(df_tab)} processos | Consulta DataJud/CNJ: {agora}"
        c.font = Font(name="Arial", color=GOLD, size=9, italic=True)
        c.fill = fl(NAVY); c.alignment = al("center"); ws.row_dimensions[2].height = 14
        ROW_H = 4; ws.row_dimensions[ROW_H].height = 30
        for ci, (col, w) in enumerate(zip(colunas, larguras), 1):
            c = ws.cell(row=ROW_H, column=ci); c.value = col
            c.font = Font(name="Arial", bold=True, color=WHITE, size=8)
            c.fill = fl(NAVY); c.alignment = al("center", wrap=True); c.border = bd()
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A{ROW_H}:{get_column_letter(len(colunas))}{ROW_H}"
        for i, (_, row) in enumerate(df_tab.iterrows()):
            er = ROW_H + 1 + i
            arq = str(row.get("ARQUIVADO DEFINITIVAMENTE", ""))
            bg = ("FFE8E8" if i % 2 == 0 else "FFD5D5") if "SIM" in arq else \
                 ("E8F5E9" if i % 2 == 0 else "D5EED8") if "NÃO" in arq else \
                 ("FFF8E8" if i % 2 == 0 else "FFF0D0")
            for ci, col in enumerate(colunas, 1):
                c = ws.cell(row=er, column=ci); c.value = row[col]
                c.border = bd(); c.fill = fl(bg); c.font = ft(size=8)
                c.alignment = al("center" if ci <= 3 else "left")
                if col == "ARQUIVADO DEFINITIVAMENTE":
                    cor_f = RED if "SIM" in arq else (GREEN if "NÃO" in arq else AMBER)
                    c.font = Font(name="Arial", bold=True, size=8, color=cor_f)
                    c.alignment = al("center")
                if col == "Status":
                    cor_s = {"ARQUIVADO": RED, "EM ANDAMENTO": GREEN, "NÃO ENCONTRADO": AMBER,
                             "ACESSO NEGADO": AMBER, "SEM CNJ PADRÃO": BLUE}.get(str(row[col]), AMBER)
                    c.font = Font(name="Arial", bold=True, size=8, color=cor_s)
            ws.row_dimensions[er].height = 15

    ws1 = wb.active; ws1.title = "Todos os Resultados"
    escrever_aba(ws1, "SEBRAE CONTENCIOSO — RESULTADO COMPLETO CONSULTA DATAJUD/CNJ", df_res, NAVY)

    ws2 = wb.create_sheet("✅ Arquivados Definitivamente")
    if len(arquivados):
        escrever_aba(ws2, "PROCESSOS ARQUIVADOS DEFINITIVAMENTE", arquivados, RED)
    else:
        ws2["A1"] = "Nenhum processo arquivado definitivamente identificado."
        ws2["A1"].font = ft(bold=True)

    ws3 = wb.create_sheet("❌ Em Andamento")
    escrever_aba(ws3, "PROCESSOS EM ANDAMENTO CONFIRMADOS", em_andamento, GREEN)

    ws4 = wb.create_sheet("❓ Verificar Manualmente")
    if len(verificar):
        escrever_aba(ws4, "STATUS INDEFINIDO — VERIFICAR MANUALMENTE NOS PORTAIS DOS TRIBUNAIS", verificar, AMBER)

    ws5 = wb.create_sheet("📊 Resumo")
    ws5.sheet_view.showGridLines = False
    ws5.merge_cells("A1:C1")
    c = ws5["A1"]; c.value = "RESUMO — CONSULTA DATAJUD CNJ"
    c.font = Font(name="Arial", bold=True, color=WHITE, size=13)
    c.fill = fl(NAVY); c.alignment = al("center"); ws5.row_dimensions[1].height = 28
    resumo = [
        ("Total consultados", len(df_res), NAVY),
        ("✅ Arquivados definitivamente", len(arquivados), RED),
        ("❌ Em andamento (confirmados)", len(em_andamento), GREEN),
        ("❓ Verificar manualmente", len(verificar), AMBER),
        ("Data/hora da consulta", agora, BLUE),
        ("Fonte", "API DataJud/CNJ", BLUE),
    ]
    for ri, (desc, val, cor) in enumerate(resumo):
        er = 3 + ri; ws5.row_dimensions[er].height = 24
        ws5.merge_cells(f"A{er}:B{er}")
        c1 = ws5.cell(row=er, column=1); c1.value = desc
        c1.font = ft(bold=True, size=10); c1.fill = fl(GRAY_L); c1.border = bd(); c1.alignment = al("left")
        c2 = ws5.cell(row=er, column=3); c2.value = val
        c2.font = Font(name="Arial", bold=True, size=12, color=cor)
        c2.fill = fl(GRAY_L); c2.border = bd(); c2.alignment = al("center")
    for ci, w in enumerate([32, 8, 18], 1):
        ws5.column_dimensions[get_column_letter(ci)].width = w

    ws5.merge_cells("A10:C10")
    c = ws5["A10"]; c.value = "PORTAIS PARA CONSULTA MANUAL (processos sem CNJ padrão)"
    c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c.fill = fl(BLUE); c.alignment = al("center"); ws5.row_dimensions[10].height = 18
    portais = [
        ("AD / DF", "TJDFT / TRF1", "https://www.tjdft.jus.br  |  https://processual.trf1.jus.br"),
        ("SP", "TJSP / TRF3", "https://esaj.tjsp.jus.br  |  https://www.trf3.jus.br"),
        ("RJ", "TJRJ / TRF2", "https://www3.tjrj.jus.br  |  https://eproc.trf2.jus.br"),
        ("MG", "TJMG / TRF1", "https://www.tjmg.jus.br  |  https://processual.trf1.jus.br"),
        ("PR / SC / RS", "TRF4", "https://eproc.trf4.jus.br"),
        ("BA / AM / GO / ES / MS", "TRF1", "https://processual.trf1.jus.br"),
    ]
    for ri, (uf, trib, url) in enumerate(portais):
        er = 11 + ri
        for ci, val in enumerate([uf, trib, url], 1):
            c = ws5.cell(row=er, column=ci); c.value = val
            c.font = ft(size=9); c.border = bd()
            c.fill = fl(GRAY_L if ri % 2 == 0 else WHITE); c.alignment = al("left")
        ws5.row_dimensions[er].height = 15

    wb.save(saida_path)


# ══════════════════════════════════════════════════════════════════════════
#  INTERFACE GRÁFICA (janela)
# ══════════════════════════════════════════════════════════════════════════
class App(tk.Tk):
    NAVY = "#1A2B5C"
    BLUE = "#2E5AAC"
    GOLD = "#D4A017"
    BG = "#F4F6FB"

    def __init__(self):
        super().__init__()
        self.title("Robô de Consulta Processual — SEBRAE Contencioso")
        self.geometry("720x580")
        self.minsize(680, 540)
        self.configure(bg=self.BG)

        self.csv_path = None
        self.headers = None
        self._parar = False
        self._rodando = False
        self._resultado = None

        self._montar_ui()
        # Prepara a chave em segundo plano assim que a janela abre
        threading.Thread(target=self._preparar_chave, daemon=True).start()

    # ---------- construção da interface ----------
    def _montar_ui(self):
        cab = tk.Frame(self, bg=self.NAVY, height=70)
        cab.pack(fill="x")
        tk.Label(cab, text="⚖️  Consulta Processual — SEBRAE Contencioso",
                 bg=self.NAVY, fg="white", font=("Segoe UI", 15, "bold")).pack(pady=6)
        tk.Label(cab, text="Verifica na API do CNJ (DataJud) quais processos estão arquivados",
                 bg=self.NAVY, fg="#C9D4EF", font=("Segoe UI", 9)).pack()

        corpo = tk.Frame(self, bg=self.BG)
        corpo.pack(fill="both", expand=True, padx=16, pady=12)

        # Passo 1 — escolher CSV
        f1 = tk.LabelFrame(corpo, text=" Passo 1 — Escolher a planilha (CSV) ",
                           bg=self.BG, fg=self.NAVY, font=("Segoe UI", 10, "bold"))
        f1.pack(fill="x", pady=(0, 10))
        linha = tk.Frame(f1, bg=self.BG); linha.pack(fill="x", padx=10, pady=10)
        self.lbl_arquivo = tk.Label(linha, text="Nenhum arquivo selecionado",
                                    bg="white", fg="#555", anchor="w",
                                    relief="solid", bd=1, font=("Segoe UI", 9), padx=8)
        self.lbl_arquivo.pack(side="left", fill="x", expand=True, ipady=5)
        tk.Button(linha, text="📂  Escolher CSV...", command=self._escolher_csv,
                  bg=self.BLUE, fg="white", font=("Segoe UI", 10, "bold"),
                  relief="flat", padx=12, cursor="hand2").pack(side="left", padx=(8, 0))

        # Passo 2 — iniciar
        f2 = tk.LabelFrame(corpo, text=" Passo 2 — Rodar a consulta ",
                           bg=self.BG, fg=self.NAVY, font=("Segoe UI", 10, "bold"))
        f2.pack(fill="x", pady=(0, 10))
        botoes = tk.Frame(f2, bg=self.BG); botoes.pack(fill="x", padx=10, pady=10)
        self.btn_iniciar = tk.Button(botoes, text="▶  Iniciar consulta", command=self._iniciar,
                                     bg="#1A7A4A", fg="white", font=("Segoe UI", 11, "bold"),
                                     relief="flat", padx=16, pady=4, cursor="hand2", state="disabled")
        self.btn_iniciar.pack(side="left")
        self.btn_parar = tk.Button(botoes, text="■  Parar", command=self._parar_consulta,
                                   bg="#C0392B", fg="white", font=("Segoe UI", 10, "bold"),
                                   relief="flat", padx=12, pady=4, cursor="hand2", state="disabled")
        self.btn_parar.pack(side="left", padx=(8, 0))
        self.btn_abrir = tk.Button(botoes, text="📊  Abrir resultado", command=self._abrir_resultado,
                                   bg=self.GOLD, fg="white", font=("Segoe UI", 10, "bold"),
                                   relief="flat", padx=12, pady=4, cursor="hand2", state="disabled")
        self.btn_abrir.pack(side="right")

        # Progresso
        self.barra = ttk.Progressbar(corpo, mode="determinate")
        self.barra.pack(fill="x", pady=(2, 4))
        self.lbl_status = tk.Label(corpo, text="Preparando a chave de acesso à API...",
                                   bg=self.BG, fg=self.NAVY, font=("Segoe UI", 10, "bold"), anchor="w")
        self.lbl_status.pack(fill="x")
        self.lbl_contadores = tk.Label(corpo, text="", bg=self.BG, fg="#333",
                                       font=("Segoe UI", 9), anchor="w")
        self.lbl_contadores.pack(fill="x", pady=(0, 6))

        # Log
        flog = tk.LabelFrame(corpo, text=" Andamento ", bg=self.BG, fg=self.NAVY,
                             font=("Segoe UI", 9, "bold"))
        flog.pack(fill="both", expand=True)
        self.txt_log = tk.Text(flog, height=8, bg="#0F1626", fg="#D6E0F5",
                               font=("Consolas", 9), relief="flat", wrap="word")
        self.txt_log.pack(fill="both", expand=True, padx=6, pady=6)
        self.txt_log.configure(state="disabled")

    # ---------- utilidades de interface ----------
    def log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        def _ins():
            self.txt_log.configure(state="normal")
            self.txt_log.insert("end", f"[{ts}] {msg}\n")
            self.txt_log.see("end")
            self.txt_log.configure(state="disabled")
        self.after(0, _ins)

    def _set_status(self, txt):
        self.after(0, lambda: self.lbl_status.config(text=txt))

    # ---------- chave ----------
    def _preparar_chave(self):
        chave = obter_chave_api(self.log)
        ativa = validar_chave(chave, self.log)
        if not ativa and chave != CHAVE_FALLBACK:
            self.log("Tentando a última chave conhecida como reserva...")
            chave = CHAVE_FALLBACK
            ativa = validar_chave(chave, self.log)
        self.headers = {"Authorization": f"APIKey {chave}", "Content-Type": "application/json"}
        if ativa:
            self._set_status("✅ Chave da API ativa. Escolha o CSV e clique em Iniciar.")
        else:
            self._set_status("⚠️ Não foi possível validar a chave — verifique a internet.")
        self._atualizar_botao_iniciar()

    def _atualizar_botao_iniciar(self):
        def _upd():
            pronto = (self.csv_path is not None) and (self.headers is not None) and not self._rodando
            self.btn_iniciar.config(state="normal" if pronto else "disabled")
        self.after(0, _upd)

    # ---------- ações ----------
    def _escolher_csv(self):
        caminho = filedialog.askopenfilename(
            title="Selecione a planilha de processos (CSV)",
            filetypes=[("Planilha CSV", "*.csv"), ("Todos os arquivos", "*.*")])
        if caminho:
            self.csv_path = caminho
            self.lbl_arquivo.config(text=Path(caminho).name, fg="#111")
            self.log(f"CSV selecionado: {Path(caminho).name}")
            self._atualizar_botao_iniciar()

    def _iniciar(self):
        if not self.csv_path or not self.headers:
            messagebox.showwarning("Atenção", "Escolha um arquivo CSV primeiro.")
            return
        self._parar = False
        self._rodando = True
        self._resultado = None
        self.btn_iniciar.config(state="disabled")
        self.btn_abrir.config(state="disabled")
        self.btn_parar.config(state="normal")
        self.barra.config(value=0)
        threading.Thread(target=self._rodar, daemon=True).start()

    def _rodar(self):
        try:
            resultado = executar_consulta(
                self.csv_path, self.headers,
                progresso_cb=self._on_progresso,
                log_cb=self.log,
                parar_flag=lambda: self._parar)
            self._resultado = resultado
            if resultado:
                self._set_status("✅ Consulta concluída! Clique em 'Abrir resultado'.")
                self.after(0, lambda: self.btn_abrir.config(state="normal"))
                self.after(0, lambda: messagebox.showinfo(
                    "Concluído",
                    "Consulta finalizada com sucesso!\n\n"
                    f"Arquivo gerado:\n{Path(resultado).name}\n\n"
                    "Ele foi salvo na mesma pasta do CSV."))
            elif self._parar:
                self._set_status("⏹ Consulta cancelada.")
            else:
                self._set_status("Nenhum resultado gerado.")
        except Exception as e:
            self.log(f"ERRO: {e}")
            self._set_status("❌ Ocorreu um erro — veja o andamento abaixo.")
            self.after(0, lambda: messagebox.showerror("Erro", f"Ocorreu um erro:\n\n{e}"))
        finally:
            self._rodando = False
            self.after(0, lambda: self.btn_parar.config(state="disabled"))
            self._atualizar_botao_iniciar()

    def _on_progresso(self, feitos, total, cont):
        def _upd():
            self.barra.config(maximum=max(total, 1), value=feitos)
            pct = int(feitos * 100 / total) if total else 0
            self.lbl_status.config(text=f"Consultando... {feitos}/{total}  ({pct}%)")
            self.lbl_contadores.config(
                text=f"✅ Arquivados: {cont['ARQUIVADO']}    "
                     f"❌ Em andamento: {cont['EM ANDAMENTO']}    "
                     f"❓ Não encontrados: {cont['NÃO ENCONTRADO']}    "
                     f"⚠️ Erros/negados: {cont['ACESSO NEGADO'] + cont['ERRO']}")
        self.after(0, _upd)

    def _parar_consulta(self):
        self._parar = True
        self.btn_parar.config(state="disabled")
        self._set_status("Parando após o processo atual...")

    def _abrir_resultado(self):
        if self._resultado and Path(self._resultado).exists():
            webbrowser.open(Path(self._resultado).as_uri())


if __name__ == "__main__":
    App().mainloop()
