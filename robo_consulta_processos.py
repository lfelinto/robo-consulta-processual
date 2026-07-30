"""
╔══════════════════════════════════════════════════════════════════════════════╗
║   ROBÔ DE CONSULTA PROCESSUAL — SEBRAE CONTENCIOSO                         ║
║   API DataJud CNJ — Elasticsearch por tribunal                             ║
║                                                                              ║
║   REQUISITOS:  pip install requests pandas openpyxl tqdm                    ║
║   EXECUÇÃO:    python robo_consulta_processos.py                            ║
║   RESULTADO:   Resultado_Consulta_Processual.xlsx                           ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import requests, pandas as pd, time, json, re, sys
from datetime import datetime
from pathlib import Path
from tqdm import tqdm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configurações ──────────────────────────────────────────────────────────
ARQUIVO_ENTRADA = "processos_para_consulta_v2.csv"
ARQUIVO_SAIDA   = "Resultado_Consulta_Processual.xlsx"
ARQUIVO_LOG     = "robo_log.txt"

API_BASE    = "https://api-publica.datajud.cnj.jus.br"
API_KEY     = "cDZHYzlZa0JadVREZDJCendFbGthcTM="   # chave pública CNJ
HEADERS     = {
    "Authorization": f"APIKey {API_KEY}",
    "Content-Type": "application/json"
}
TIMEOUT     = 15
PAUSA       = 0.65    # ~92 req/min, abaixo do limite de 100
PAUSA_429   = 10.0
MAX_TENT    = 3

# Palavras que indicam arquivamento definitivo (busca no texto JSON completo)
KW_ARQUIVADO = [
    "arquivado definitivamente", "arquivamento definitivo",
    "baixado definitivamente", "baixa definitiva",
    "transitado em julgado", "trânsito em julgado",
    "processo encerrado", "julgamento definitivo",
    "extinção definitiva", "sentença definitiva",
    "encerrado definitivamente", "baixa dos autos",
]

# Mapeamento de número CNJ → índice DataJud
INDICE_MAP = {
    "4.01":"trf1","4.02":"trf2","4.03":"trf3","4.04":"trf4",
    "4.05":"trf5","4.06":"trf6",
    "5.01":"trt1","5.02":"trt2","5.03":"trt3","5.04":"trt4",
    "5.05":"trt5","5.06":"trt6","5.07":"trt7","5.08":"trt8",
    "5.09":"trt9","5.10":"trt10","5.11":"trt11","5.12":"trt12",
    "5.13":"trt13","5.14":"trt14","5.15":"trt15","5.17":"trt17",
    "5.18":"trt18","5.20":"trt20","5.21":"trt21","5.22":"trt22",
    "5.23":"trt23","5.24":"trt24",
    "8.01":"tjac","8.02":"tjal","8.03":"tjap","8.04":"tjam",
    "8.05":"tjba","8.06":"tjce","8.07":"tjdft","8.08":"tjes",
    "8.09":"tjgo","8.10":"tjma","8.11":"tjmt","8.12":"tjms",
    "8.13":"tjmg","8.14":"tjpa","8.15":"tjpb","8.16":"tjpr",
    "8.17":"tjpe","8.18":"tjpi","8.19":"tjrj","8.20":"tjrn",
    "8.21":"tjrs","8.22":"tjro","8.23":"tjrr","8.24":"tjsc",
    "8.25":"tjsp","8.26":"tjsp","8.27":"tjse","8.28":"tjto",
    "1.00":"stf","2.00":"stj","3.00":"tst",
}

# ── Helpers ────────────────────────────────────────────────────────────────
def log(msg, nivel="INFO"):
    ts = datetime.now().strftime("%H:%M:%S")
    linha = f"[{ts}][{nivel}] {msg}"
    print(linha)
    with open(ARQUIVO_LOG, "a", encoding="utf-8") as f:
        f.write(linha + "\n")

def indice_datajud(cnj: str) -> str | None:
    """Extrai o índice DataJud do número CNJ."""
    m = re.match(r"\d{7}-\d{2}\.\d{4}\.(\d)\.(\d{2})\.", cnj)
    if not m: return None
    cod = f"{m.group(1)}.{m.group(2)}"
    nome = INDICE_MAP.get(cod)
    return f"api_publica_{nome}" if nome else None

def is_arquivado(dados: dict) -> tuple[bool, str]:
    """Analisa JSON retornado e detecta arquivamento definitivo."""
    texto = json.dumps(dados, ensure_ascii=False).lower()
    for kw in KW_ARQUIVADO:
        if kw in texto:
            return True, f"Palavra-chave: «{kw}»"

    # Verificar último movimento explicitamente
    movs = dados.get("movimentos", [])
    if movs:
        ult = movs[-1]
        desc = json.dumps(ult, ensure_ascii=False).lower()
        for kw in KW_ARQUIVADO:
            if kw in desc:
                return True, f"Último movimento: «{kw}»"

    # Verificar campo de classe / situação
    for campo in ["classeProcessual","situacao","grau","fase"]:
        val = str(dados.get(campo,"")).lower()
        for kw in KW_ARQUIVADO:
            if kw in val:
                return True, f"Campo '{campo}': {val[:60]}"

    return False, "Sem indicação de arquivamento"

def extrair_info(dados: dict) -> dict:
    """Extrai campos úteis do JSON do processo."""
    movs = dados.get("movimentos", [])
    ult_data = "—"
    ult_mov  = "—"
    if movs:
        ult = movs[-1]
        ult_data = ult.get("dataHora", ult.get("data","—"))
        if isinstance(ult_data, str) and len(ult_data) >= 10:
            ult_data = ult_data[:10]
        cod = ult.get("codigo","")
        ult_mov = ult.get("nome", ult.get("descricao", str(cod)))[:80]

    partes = dados.get("partes", [])
    polo_p = next((p.get("nome","") for p in partes
                   if str(p.get("polo","")).upper() in ("PASSIVO","REU","RÉU")), "—")

    return {
        "Tribunal":           dados.get("tribunal",{}).get("nome", dados.get("orgaoJulgador",{}).get("nome","—"))[:50],
        "Classe":             dados.get("classeProcessual",{}).get("nome","—")[:40],
        "Grau":               dados.get("grau","—"),
        "Data Última Mov.":   ult_data,
        "Última Movimentação":ult_mov,
        "Réu/Passivo":        polo_p[:60],
        "Qtd Movimentos":     len(movs),
    }

def consultar(cnj: str, indice: str, tentativa: int = 1) -> tuple[str, str, dict]:
    """
    Consulta DataJud via Elasticsearch.
    Retorna (status, descricao, dados_extraidos)
    """
    url = f"{API_BASE}/{indice}/_search"
    body = {
        "size": 1,
        "query": {"match": {"numeroProcesso": cnj}},
        "_source": ["numeroProcesso","classeProcessual","grau","tribunal",
                    "orgaoJulgador","movimentos","partes","situacao"]
    }
    try:
        r = requests.post(url, headers=HEADERS, json=body, timeout=TIMEOUT)

        if r.status_code == 200:
            hits = r.json().get("hits", {}).get("hits", [])
            if not hits:
                return "NÃO ENCONTRADO", "Processo não localizado neste tribunal no DataJud", {}
            dados = hits[0]["_source"]
            arquivado, motivo = is_arquivado(dados)
            info = extrair_info(dados)
            status = "ARQUIVADO" if arquivado else "EM ANDAMENTO"
            return status, motivo, info

        elif r.status_code == 404:
            return "NÃO ENCONTRADO", "Índice/tribunal não disponível no DataJud", {}

        elif r.status_code == 429:
            log(f"Rate limit. Aguardando {PAUSA_429}s...", "AVISO")
            time.sleep(PAUSA_429)
            if tentativa < MAX_TENT:
                return consultar(cnj, indice, tentativa + 1)
            return "ERRO 429", "Rate limit excedido após retentativas", {}

        elif r.status_code == 403:
            return "ACESSO NEGADO", "Tribunal não expõe dados via API pública DataJud", {}

        else:
            if tentativa < MAX_TENT:
                time.sleep(2)
                return consultar(cnj, indice, tentativa + 1)
            return f"ERRO HTTP {r.status_code}", r.text[:100], {}

    except requests.ConnectionError:
        return "ERRO CONEXÃO", "Sem acesso à internet ou API indisponível", {}
    except requests.Timeout:
        if tentativa < MAX_TENT:
            return consultar(cnj, indice, tentativa + 1)
        return "TIMEOUT", "Servidor não respondeu em 15s", {}
    except Exception as e:
        return "ERRO", str(e)[:100], {}

# ── EXECUÇÃO PRINCIPAL ─────────────────────────────────────────────────────
def main():
    inicio = datetime.now()
    log(f"══ ROBÔ CONSULTA PROCESSUAL SEBRAE — {inicio.strftime('%d/%m/%Y %H:%M')} ══")

    if not Path(ARQUIVO_ENTRADA).exists():
        log(f"Arquivo '{ARQUIVO_ENTRADA}' não encontrado.", "ERRO")
        sys.exit(1)

    df = pd.read_csv(ARQUIVO_ENTRADA, encoding="utf-8-sig")
    log(f"Carregados {len(df)} processos.")

    # Garantir coluna de índice DataJud
    if "Indice_DataJud" not in df.columns or df["Indice_DataJud"].isna().all():
        df["Indice_DataJud"] = df["CNJ_Extraido"].apply(
            lambda x: f"api_publica_{INDICE_MAP.get(re.match(r'\\d{7}-\\d{2}\\.\\d{4}\\.(\\d)\\.(\\d{2})\\.', str(x) or '').group(1)+'.'+re.match(r'\\d{7}-\\d{2}\\.\\d{4}\\.(\\d)\\.(\\d{2})\\.', str(x) or '').group(2),'')}"
            if pd.notna(x) else None
        )

    resultados = []
    sem_indice = []

    for _, row in df.iterrows():
        if pd.isna(row.get("CNJ_Extraido")) or str(row.get("CNJ_Extraido","")).strip() == "":
            sem_indice.append(row)
        elif pd.isna(row.get("Indice_DataJud")) or "nan" in str(row.get("Indice_DataJud","")):
            sem_indice.append(row)
        else:
            resultados.append(row)

    com_cnj  = pd.DataFrame(resultados)
    sem_cnj2 = pd.DataFrame(sem_indice)
    log(f"  → Consultáveis via API: {len(com_cnj)}")
    log(f"  → Sem índice/CNJ válido: {len(sem_cnj2)}")
    log(f"  → Tempo estimado: ~{int(len(com_cnj)*PAUSA/60)+1} minutos")

    saida = []

    # Contadores por status (para checkpoint)
    contadores = {"ARQUIVADO":0, "EM ANDAMENTO":0, "NÃO ENCONTRADO":0,
                  "ACESSO NEGADO":0, "ERRO":0}

    for i, (_, row) in enumerate(tqdm(com_cnj.iterrows(), total=len(com_cnj),
                                       desc="Consultando CNJ", unit="proc")):
        cnj    = str(row["CNJ_Extraido"]).strip()
        indice = str(row["Indice_DataJud"]).strip()
        ctrl   = str(row.get("Controle Interno do Sebrae",""))
        estado = str(row.get("Estado",""))

        status, descricao, info = consultar(cnj, indice)

        # Atualizar contadores
        chave = status if status in contadores else "ERRO"
        contadores[chave] = contadores.get(chave, 0) + 1

        saida.append({
            "#":                          i + 1,
            "Controle Interno":           ctrl,
            "Estado":                     estado,
            "CNJ":                        cnj,
            "Número Original":            str(row.get("Número do Processo","")),
            "Classe de Ação":             str(row.get("Classe de Ação","")),
            "Índice DataJud":             indice,
            "Status":                     status,
            "Detalhamento":               descricao[:120],
            "Tribunal":                   info.get("Tribunal","—"),
            "Classe Processo":            info.get("Classe","—"),
            "Grau":                       info.get("Grau","—"),
            "Data Última Movimentação":   info.get("Data Última Mov.","—"),
            "Última Movimentação":        info.get("Última Movimentação","—"),
            "Qtd. Movimentos":            info.get("Qtd Movimentos","—"),
            "ARQUIVADO DEFINITIVAMENTE":  "✅ SIM" if status=="ARQUIVADO" else (
                                          "❌ NÃO" if status=="EM ANDAMENTO" else "❓ VERIFICAR"),
        })

        # Checkpoint a cada 50
        if (i + 1) % 50 == 0:
            parcial = pd.DataFrame(saida)
            parcial.to_csv("checkpoint_consulta.csv", index=False, encoding="utf-8-sig")
            log(f"Checkpoint {i+1}/{len(com_cnj)} | "
                f"Arquivados:{contadores['ARQUIVADO']} | "
                f"Andamento:{contadores['EM ANDAMENTO']} | "
                f"Não encontrado:{contadores['NÃO ENCONTRADO']}")

        time.sleep(PAUSA)

    # Registrar sem CNJ padrão
    for j, (_, row) in enumerate(sem_cnj2.iterrows()):
        saida.append({
            "#":                         len(com_cnj) + j + 1,
            "Controle Interno":          str(row.get("Controle Interno do Sebrae","")),
            "Estado":                    str(row.get("Estado","")),
            "CNJ":                       "—",
            "Número Original":           str(row.get("Número do Processo","")),
            "Classe de Ação":            str(row.get("Classe de Ação","")),
            "Índice DataJud":            "—",
            "Status":                    "SEM CNJ PADRÃO",
            "Detalhamento":              "Número em formato antigo — consultar portal do tribunal",
            "Tribunal":"—","Classe Processo":"—","Grau":"—",
            "Data Última Movimentação":  "—",
            "Última Movimentação":       "—",
            "Qtd. Movimentos":           "—",
            "ARQUIVADO DEFINITIVAMENTE": "❓ VERIFICAR MANUAL",
        })

    df_res = pd.DataFrame(saida)

    # Subsets por status
    arquivados   = df_res[df_res["ARQUIVADO DEFINITIVAMENTE"] == "✅ SIM"]
    em_andamento = df_res[df_res["ARQUIVADO DEFINITIVAMENTE"] == "❌ NÃO"]
    verificar    = df_res[~df_res["ARQUIVADO DEFINITIVAMENTE"].isin(["✅ SIM","❌ NÃO"])]

    fim = datetime.now()
    duracao = int((fim - inicio).total_seconds())
    log(f"\n══ RESULTADO FINAL ══")
    log(f"Total consultados:           {len(df_res)}")
    log(f"Arquivados definitivamente:  {len(arquivados)}")
    log(f"Em andamento:                {len(em_andamento)}")
    log(f"Verificar manualmente:       {len(verificar)}")
    log(f"Duração total:               {duracao//60}min {duracao%60}s")

    salvar_xlsx(df_res, arquivados, em_andamento, verificar, fim)
    log(f"\nArquivo gerado: {ARQUIVO_SAIDA}")

# ── SALVAR XLSX ────────────────────────────────────────────────────────────
def salvar_xlsx(df_res, arquivados, em_andamento, verificar, fim):
    NAVY="1A2B5C"; BLUE="2E5AAC"; RED="C0392B"; GREEN="1A7A4A"
    AMBER="D97706"; WHITE="FFFFFF"; GRAY_L="F4F6FB"; GOLD="D4A017"

    def fl(c): return PatternFill("solid",fgColor=c)
    def ft(bold=False, color="1A2B5C", size=9):
        return Font(name="Arial", bold=bold, color=color, size=size)
    def al(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def bd():
        s=Side(border_style="thin",color="C5CDE8")
        return Border(left=s,right=s,top=s,bottom=s)

    wb = openpyxl.Workbook()
    agora = fim.strftime("%d/%m/%Y %H:%M")

    colunas   = df_res.columns.tolist()
    larguras  = [4,18,6,26,34,22,22,18,50,30,28,6,18,40,14,22]
    larguras += [15]*(len(colunas)-len(larguras))

    def escrever_aba(ws, titulo, df_tab, cor):
        ws.sheet_view.showGridLines = False
        ws.merge_cells(f"A1:{get_column_letter(len(colunas))}1")
        c=ws["A1"]; c.value=titulo
        c.font=Font(name="Arial",bold=True,color=WHITE,size=12)
        c.fill=fl(cor); c.alignment=al("center"); ws.row_dimensions[1].height=28

        ws.merge_cells(f"A2:{get_column_letter(len(colunas))}2")
        c=ws["A2"]
        c.value=f"{len(df_tab)} processos | Consulta DataJud/CNJ: {agora}"
        c.font=Font(name="Arial",color=GOLD,size=9,italic=True)
        c.fill=fl(NAVY); c.alignment=al("center"); ws.row_dimensions[2].height=14

        ROW_H=4; ws.row_dimensions[ROW_H].height=30
        for ci,(col,w) in enumerate(zip(colunas,larguras),1):
            c=ws.cell(row=ROW_H,column=ci); c.value=col
            c.font=Font(name="Arial",bold=True,color=WHITE,size=8)
            c.fill=fl(NAVY); c.alignment=al("center",wrap=True); c.border=bd()
            ws.column_dimensions[get_column_letter(ci)].width=w

        ws.freeze_panes="A5"
        ws.auto_filter.ref=f"A{ROW_H}:{get_column_letter(len(colunas))}{ROW_H}"

        for i,(_, row) in enumerate(df_tab.iterrows()):
            er=ROW_H+1+i
            arq=str(row.get("ARQUIVADO DEFINITIVAMENTE",""))
            bg=("FFE8E8" if i%2==0 else "FFD5D5") if "SIM" in arq else \
               ("E8F5E9" if i%2==0 else "D5EED8") if "NÃO" in arq else \
               ("FFF8E8" if i%2==0 else "FFF0D0")
            for ci,col in enumerate(colunas,1):
                c=ws.cell(row=er,column=ci); c.value=row[col]
                c.border=bd(); c.fill=fl(bg); c.font=ft(size=8)
                c.alignment=al("center" if ci<=3 else "left")
                if col=="ARQUIVADO DEFINITIVAMENTE":
                    cor_f=RED if "SIM" in arq else (GREEN if "NÃO" in arq else AMBER)
                    c.font=Font(name="Arial",bold=True,size=8,color=cor_f)
                    c.alignment=al("center")
                if col=="Status":
                    cor_s={"ARQUIVADO":RED,"EM ANDAMENTO":GREEN,"NÃO ENCONTRADO":AMBER,
                           "ACESSO NEGADO":AMBER,"SEM CNJ PADRÃO":BLUE}.get(str(row[col]),AMBER)
                    c.font=Font(name="Arial",bold=True,size=8,color=cor_s)
            ws.row_dimensions[er].height=15

    # Aba 1: Todos
    ws1=wb.active; ws1.title="Todos os Resultados"
    escrever_aba(ws1,"SEBRAE CONTENCIOSO — RESULTADO COMPLETO CONSULTA DATAJUD/CNJ",df_res,NAVY)

    # Aba 2: Arquivados
    ws2=wb.create_sheet("✅ Arquivados Definitivamente")
    if len(arquivados):
        escrever_aba(ws2,"PROCESSOS ARQUIVADOS DEFINITIVAMENTE",arquivados,RED)
    else:
        ws2["A1"]="Nenhum processo arquivado definitivamente identificado."
        ws2["A1"].font=ft(bold=True)

    # Aba 3: Em andamento
    ws3=wb.create_sheet("❌ Em Andamento")
    escrever_aba(ws3,"PROCESSOS EM ANDAMENTO CONFIRMADOS",em_andamento,GREEN)

    # Aba 4: Verificar
    ws4=wb.create_sheet("❓ Verificar Manualmente")
    if len(verificar):
        escrever_aba(ws4,"STATUS INDEFINIDO — VERIFICAR MANUALMENTE NOS PORTAIS DOS TRIBUNAIS",verificar,AMBER)

    # Aba 5: Resumo
    ws5=wb.create_sheet("📊 Resumo")
    ws5.sheet_view.showGridLines=False
    ws5.merge_cells("A1:C1")
    c=ws5["A1"]; c.value="RESUMO — CONSULTA DATAJUD CNJ"
    c.font=Font(name="Arial",bold=True,color=WHITE,size=13)
    c.fill=fl(NAVY); c.alignment=al("center"); ws5.row_dimensions[1].height=28

    resumo=[
        ("Total consultados",                  len(df_res),       NAVY),
        ("✅ Arquivados definitivamente",       len(arquivados),   RED),
        ("❌ Em andamento (confirmados)",       len(em_andamento), GREEN),
        ("❓ Verificar manualmente",            len(verificar),    AMBER),
        ("Data/hora da consulta",              agora,             BLUE),
        ("Fonte",                              "API DataJud/CNJ", BLUE),
    ]
    for ri,(desc,val,cor) in enumerate(resumo):
        er=3+ri; ws5.row_dimensions[er].height=24
        ws5.merge_cells(f"A{er}:B{er}")
        c1=ws5.cell(row=er,column=1); c1.value=desc
        c1.font=ft(bold=True,size=10); c1.fill=fl(GRAY_L); c1.border=bd(); c1.alignment=al("left")
        c2=ws5.cell(row=er,column=3); c2.value=val
        c2.font=Font(name="Arial",bold=True,size=12,color=cor)
        c2.fill=fl(GRAY_L); c2.border=bd(); c2.alignment=al("center")

    for ci,w in enumerate([32,8,18],1):
        ws5.column_dimensions[get_column_letter(ci)].width=w

    # Orientações portais por estado
    ws5.merge_cells("A10:C10")
    c=ws5["A10"]; c.value="PORTAIS PARA CONSULTA MANUAL (processos sem CNJ padrão)"
    c.font=Font(name="Arial",bold=True,color=WHITE,size=10)
    c.fill=fl(BLUE); c.alignment=al("center"); ws5.row_dimensions[10].height=18

    portais=[
        ("AD / DF","TJDFT / TRF1","https://www.tjdft.jus.br  |  https://processual.trf1.jus.br"),
        ("SP","TJSP / TRF3","https://esaj.tjsp.jus.br  |  https://www.trf3.jus.br"),
        ("RJ","TJRJ / TRF2","https://www3.tjrj.jus.br  |  https://eproc.trf2.jus.br"),
        ("MG","TJMG / TRF1","https://www.tjmg.jus.br  |  https://processual.trf1.jus.br"),
        ("PR / SC / RS","TRF4","https://eproc.trf4.jus.br"),
        ("BA / AM / GO / ES / MS","TRF1","https://processual.trf1.jus.br"),
    ]
    for ri,(uf,trib,url) in enumerate(portais):
        er=11+ri
        for ci,val in enumerate([uf,trib,url],1):
            c=ws5.cell(row=er,column=ci); c.value=val
            c.font=ft(size=9); c.border=bd()
            c.fill=fl(GRAY_L if ri%2==0 else WHITE); c.alignment=al("left")
        ws5.row_dimensions[er].height=15

    wb.save(ARQUIVO_SAIDA)

if __name__ == "__main__":
    main()
