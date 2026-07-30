# -*- coding: utf-8 -*-
"""
Núcleo do Robô de Consulta Processual — SEBRAE Contencioso.

Contém APENAS a lógica de consulta (sem interface): busca de chave da API,
consulta ao DataJud/CNJ, detecção de arquivamento e geração do Excel.
É reutilizado tanto pela versão desktop quanto pelo app web (Streamlit).
"""

import re
import time
import unicodedata
from datetime import datetime
from io import BytesIO

import requests
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configurações ──────────────────────────────────────────────────────────
API_BASE = "https://api-publica.datajud.cnj.jus.br"

# Página oficial do CNJ de onde a chave pública é lida automaticamente:
URL_CHAVE_CNJ = "https://datajud-wiki.cnj.jus.br/api-publica/acesso/"

# Última chave pública conhecida (usada só se a leitura automática falhar).
# Atualizada em 17/07/2026.
CHAVE_FALLBACK = "cDZHYzlZa0JadVREZDJCendQbXY6SkJlTzNjLV9TRENyQk1RdnFKZGRQdw=="

TIMEOUT = 15
PAUSA = 0.65      # ~92 req/min, abaixo do limite de 100
PAUSA_429 = 10.0
MAX_TENT = 3

# Códigos de movimento do CNJ (TPU) que indicam ARQUIVAMENTO DEFINITIVO.
#   246 = "Definitivo"  (Arquivamento Definitivo)
#    22 = "Baixa Definitiva"
CODIGOS_ARQUIVAMENTO_DEFINITIVO = {22, 246}

# Mapeamento de número CNJ → índice DataJud
INDICE_MAP = {
    "4.01": "trf1", "4.02": "trf2", "4.03": "trf3", "4.04": "trf4",
    "4.05": "trf5", "4.06": "trf6",
    "5.01": "trt1", "5.02": "trt2", "5.03": "trt3", "5.04": "trt4",
    "5.05": "trt5", "5.06": "trt6", "5.07": "trt7", "5.08": "trt8",
    "5.09": "trt9", "5.10": "trt10", "5.11": "trt11", "5.12": "trt12",
    "5.13": "trt13", "5.14": "trt14", "5.15": "trt15", "5.17": "trt17",
    "5.18": "trt18", "5.20": "trt20", "5.21": "trt21", "5.22": "trt22",
    "5.23": "trt23", "5.24": "trt24",
    "8.01": "tjac", "8.02": "tjal", "8.03": "tjap", "8.04": "tjam",
    "8.05": "tjba", "8.06": "tjce", "8.07": "tjdft", "8.08": "tjes",
    "8.09": "tjgo", "8.10": "tjma", "8.11": "tjmt", "8.12": "tjms",
    "8.13": "tjmg", "8.14": "tjpa", "8.15": "tjpb", "8.16": "tjpr",
    "8.17": "tjpe", "8.18": "tjpi", "8.19": "tjrj", "8.20": "tjrn",
    "8.21": "tjrs", "8.22": "tjro", "8.23": "tjrr", "8.24": "tjsc",
    "8.25": "tjsp", "8.26": "tjsp", "8.27": "tjse", "8.28": "tjto",
    "1.00": "stf", "2.00": "stj", "3.00": "tst",
}


# ── Chave de API (auto-atualização) ─────────────────────────────────────────
def obter_chave_api(logfn=print) -> str:
    """Busca a chave pública ATUAL no site do CNJ; usa a última conhecida se falhar."""
    try:
        r = requests.get(URL_CHAVE_CNJ, timeout=TIMEOUT)
        if r.status_code == 200:
            candidatos = re.findall(r"[A-Za-z0-9+/]{50,}={0,2}", r.text)
            for cand in candidatos:
                if cand.endswith("="):
                    logfn("Chave pública obtida automaticamente do site do CNJ.")
                    return cand
        logfn("Não foi possível ler a chave no site do CNJ; usando a última conhecida.")
    except Exception as e:
        logfn(f"Sem acesso ao site do CNJ ({type(e).__name__}); usando a última chave conhecida.")
    return CHAVE_FALLBACK


def validar_chave(chave: str, logfn=print) -> bool:
    """Faz uma chamada de teste para confirmar que a chave está ativa."""
    headers = {"Authorization": f"APIKey {chave}", "Content-Type": "application/json"}
    try:
        r = requests.post(f"{API_BASE}/api_publica_tjsp/_search",
                          headers=headers, json={"size": 1, "query": {"match_all": {}}},
                          timeout=TIMEOUT)
        if r.status_code == 200:
            logfn("Chave da API validada com sucesso (ativa).")
            return True
        logfn(f"Chave rejeitada pelo servidor (HTTP {r.status_code}).")
        return False
    except Exception as e:
        logfn(f"Não foi possível validar a chave agora ({type(e).__name__}).")
        return False


def preparar_headers(logfn=print) -> dict:
    """Obtém e valida a chave, com fallback, e retorna os headers prontos."""
    chave = obter_chave_api(logfn)
    ativa = validar_chave(chave, logfn)
    if not ativa and chave != CHAVE_FALLBACK:
        logfn("Tentando a última chave conhecida como reserva...")
        chave = CHAVE_FALLBACK
        validar_chave(chave, logfn)
    return {"Authorization": f"APIKey {chave}", "Content-Type": "application/json"}


# ── Lógica de consulta ──────────────────────────────────────────────────────
def indice_datajud(cnj: str):
    m = re.match(r"\d{7}-\d{2}\.\d{4}\.(\d)\.(\d{2})\.", str(cnj))
    if not m:
        return None
    nome = INDICE_MAP.get(f"{m.group(1)}.{m.group(2)}")
    return f"api_publica_{nome}" if nome else None


def _sem_acento(s) -> str:
    s = str(s or "").lower()
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))


def _ultimo_movimento(movs):
    if not movs:
        return None
    return sorted(movs, key=lambda m: str(m.get("dataHora", "")))[-1]


def is_arquivado(dados: dict):
    """Detecta arquivamento definitivo pelo ÚLTIMO movimento do processo."""
    ult = _ultimo_movimento(dados.get("movimentos", []) or [])
    if not ult:
        return False, "Processo sem movimentos registrados"
    cod = ult.get("codigo")
    try:
        cod_int = int(cod)
    except (TypeError, ValueError):
        cod_int = None
    nome = ult.get("nome", "")
    nome_n = _sem_acento(nome)
    if cod_int in CODIGOS_ARQUIVAMENTO_DEFINITIVO:
        return True, f"Último movimento: «{nome}» (cód. {cod})"
    if "baixa definitiva" in nome_n or nome_n == "definitivo" or \
       ("arquiv" in nome_n and "defin" in nome_n):
        return True, f"Último movimento: «{nome}»"
    return False, f"Último movimento: «{nome}» — sem arquivamento definitivo"


def extrair_info(dados: dict):
    movs = dados.get("movimentos", []) or []
    ult = _ultimo_movimento(movs)
    ult_data, ult_mov = "—", "—"
    if ult:
        d = ult.get("dataHora", "")
        ult_data = d[:10] if isinstance(d, str) and len(d) >= 10 else "—"
        ult_mov = str(ult.get("nome", "—"))[:80]
    return {
        "Tribunal": str(dados.get("tribunal", "—"))[:50],
        "Classe": str((dados.get("classe") or {}).get("nome", "—"))[:40],
        "Grau": dados.get("grau", "—"),
        "Data Última Mov.": ult_data,
        "Última Movimentação": ult_mov,
        "Qtd Movimentos": len(movs),
    }


def consultar(cnj, indice, headers, tentativa=1):
    url = f"{API_BASE}/{indice}/_search"
    cnj_digitos = re.sub(r"\D", "", str(cnj))
    body = {
        "size": 1,
        "query": {"term": {"numeroProcesso.keyword": cnj_digitos}},
        "_source": ["numeroProcesso", "classe", "grau", "tribunal",
                    "orgaoJulgador", "movimentos"],
    }
    try:
        r = requests.post(url, headers=headers, json=body, timeout=TIMEOUT)
        if r.status_code == 200:
            hits = r.json().get("hits", {}).get("hits", [])
            if not hits:
                return "NÃO ENCONTRADO", "Processo não localizado neste tribunal no DataJud", {}
            dados = hits[0]["_source"]
            arquivado, motivo = is_arquivado(dados)
            info = extrair_info(dados)
            return ("ARQUIVADO" if arquivado else "EM ANDAMENTO"), motivo, info
        elif r.status_code == 404:
            return "NÃO ENCONTRADO", "Índice/tribunal não disponível no DataJud", {}
        elif r.status_code == 429:
            time.sleep(PAUSA_429)
            if tentativa < MAX_TENT:
                return consultar(cnj, indice, headers, tentativa + 1)
            return "ERRO 429", "Rate limit excedido após retentativas", {}
        elif r.status_code in (401, 403):
            return "ACESSO NEGADO", f"HTTP {r.status_code} — chave inválida ou tribunal restrito", {}
        else:
            if tentativa < MAX_TENT:
                time.sleep(2)
                return consultar(cnj, indice, headers, tentativa + 1)
            return f"ERRO HTTP {r.status_code}", r.text[:100], {}
    except requests.ConnectionError:
        return "ERRO CONEXÃO", "Sem acesso à internet ou API indisponível", {}
    except requests.Timeout:
        if tentativa < MAX_TENT:
            return consultar(cnj, indice, headers, tentativa + 1)
        return "TIMEOUT", "Servidor não respondeu em 15s", {}
    except Exception as e:
        return "ERRO", str(e)[:100], {}


def ler_csv(fonte) -> pd.DataFrame:
    """Lê o CSV (caminho ou file-like), tolerando encoding/separador."""
    try:
        return pd.read_csv(fonte, encoding="utf-8-sig")
    except Exception:
        if hasattr(fonte, "seek"):
            fonte.seek(0)
        return pd.read_csv(fonte, encoding="latin-1", sep=None, engine="python")


def executar_consulta(df, headers, progresso_cb=None, log_cb=print, parar_flag=lambda: False):
    """
    Percorre o DataFrame e consulta o DataJud.
    progresso_cb(feitos, total, contadores) — chamado a cada processo (opcional).
    Retorna (df_resultado, contadores).
    """
    if "Indice_DataJud" not in df.columns or df["Indice_DataJud"].isna().all():
        df["Indice_DataJud"] = df.get("CNJ_Extraido", pd.Series([None] * len(df))).apply(indice_datajud)

    com_cnj, sem_cnj = [], []
    for _, row in df.iterrows():
        cnj = str(row.get("CNJ_Extraido", "") or "").strip()
        idx = str(row.get("Indice_DataJud", "") or "").strip()
        if not cnj or not idx or idx.lower() in ("nan", "none"):
            sem_cnj.append(row)
        else:
            com_cnj.append(row)
    com_cnj = pd.DataFrame(com_cnj)
    sem_cnj = pd.DataFrame(sem_cnj)
    log_cb(f"Consultáveis pela API: {len(com_cnj)} | Sem CNJ padrão: {len(sem_cnj)}")

    saida = []
    contadores = {"ARQUIVADO": 0, "EM ANDAMENTO": 0, "NÃO ENCONTRADO": 0,
                  "ACESSO NEGADO": 0, "ERRO": 0}
    total = len(com_cnj)

    for i, (_, row) in enumerate(com_cnj.iterrows()):
        if parar_flag():
            log_cb("Consulta cancelada pelo usuário.")
            break
        cnj = str(row["CNJ_Extraido"]).strip()
        indice = str(row["Indice_DataJud"]).strip()
        status, descricao, info = consultar(cnj, indice, headers)
        chave = status if status in contadores else "ERRO"
        contadores[chave] = contadores.get(chave, 0) + 1

        saida.append({
            "#": i + 1,
            "Controle Interno": str(row.get("Controle Interno do Sebrae", "")),
            "Estado": str(row.get("Estado", "")),
            "CNJ": cnj,
            "Número Original": str(row.get("Número do Processo", "")),
            "Classe de Ação": str(row.get("Classe de Ação", "")),
            "Índice DataJud": indice,
            "Status": status,
            "Detalhamento": descricao[:120],
            "Tribunal": info.get("Tribunal", "—"),
            "Classe Processo": info.get("Classe", "—"),
            "Grau": info.get("Grau", "—"),
            "Data Última Movimentação": info.get("Data Última Mov.", "—"),
            "Última Movimentação": info.get("Última Movimentação", "—"),
            "Qtd. Movimentos": info.get("Qtd Movimentos", "—"),
            "ARQUIVADO DEFINITIVAMENTE": "✅ SIM" if status == "ARQUIVADO" else (
                "❌ NÃO" if status == "EM ANDAMENTO" else "❓ VERIFICAR"),
        })
        if progresso_cb:
            progresso_cb(i + 1, total, contadores)
        time.sleep(PAUSA)

    for j, (_, row) in enumerate(sem_cnj.iterrows()):
        saida.append({
            "#": len(com_cnj) + j + 1,
            "Controle Interno": str(row.get("Controle Interno do Sebrae", "")),
            "Estado": str(row.get("Estado", "")),
            "CNJ": "—",
            "Número Original": str(row.get("Número do Processo", "")),
            "Classe de Ação": str(row.get("Classe de Ação", "")),
            "Índice DataJud": "—",
            "Status": "SEM CNJ PADRÃO",
            "Detalhamento": "Número em formato antigo — consultar portal do tribunal",
            "Tribunal": "—", "Classe Processo": "—", "Grau": "—",
            "Data Última Movimentação": "—", "Última Movimentação": "—",
            "Qtd. Movimentos": "—",
            "ARQUIVADO DEFINITIVAMENTE": "❓ VERIFICAR MANUAL",
        })

    return pd.DataFrame(saida), contadores


# ── SALVAR XLSX ─────────────────────────────────────────────────────────────
def gerar_xlsx(df_res, fim=None) -> bytes:
    """Gera o relatório Excel (5 abas) e retorna os bytes do arquivo."""
    fim = fim or datetime.now()
    arquivados = df_res[df_res["ARQUIVADO DEFINITIVAMENTE"] == "✅ SIM"]
    em_andamento = df_res[df_res["ARQUIVADO DEFINITIVAMENTE"] == "❌ NÃO"]
    verificar = df_res[~df_res["ARQUIVADO DEFINITIVAMENTE"].isin(["✅ SIM", "❌ NÃO"])]

    NAVY = "1A2B5C"; BLUE = "2E5AAC"; RED = "C0392B"; GREEN = "1A7A4A"
    AMBER = "D97706"; WHITE = "FFFFFF"; GRAY_L = "F4F6FB"; GOLD = "D4A017"

    def fl(c): return PatternFill("solid", fgColor=c)
    def ft(bold=False, color="1A2B5C", size=9):
        return Font(name="Arial", bold=bold, color=color, size=size)
    def al(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def bd():
        s = Side(border_style="thin", color="C5CDE8")
        return Border(left=s, right=s, top=s, bottom=s)

    wb = openpyxl.Workbook()
    agora = fim.strftime("%d/%m/%Y %H:%M")
    colunas = df_res.columns.tolist()
    larguras = [4, 18, 6, 26, 34, 22, 22, 18, 50, 30, 28, 6, 18, 40, 14, 22]
    larguras += [15] * (len(colunas) - len(larguras))

    def escrever_aba(ws, titulo, df_tab, cor):
        ws.sheet_view.showGridLines = False
        ws.merge_cells(f"A1:{get_column_letter(len(colunas))}1")
        c = ws["A1"]; c.value = titulo
        c.font = Font(name="Arial", bold=True, color=WHITE, size=12)
        c.fill = fl(cor); c.alignment = al("center"); ws.row_dimensions[1].height = 28
        ws.merge_cells(f"A2:{get_column_letter(len(colunas))}2")
        c = ws["A2"]
        c.value = f"{len(df_tab)} processos | Consulta DataJud/CNJ: {agora}"
        c.font = Font(name="Arial", color=GOLD, size=9, italic=True)
        c.fill = fl(NAVY); c.alignment = al("center"); ws.row_dimensions[2].height = 14
        ROW_H = 4; ws.row_dimensions[ROW_H].height = 30
        for ci, (col, w) in enumerate(zip(colunas, larguras), 1):
            c = ws.cell(row=ROW_H, column=ci); c.value = col
            c.font = Font(name="Arial", bold=True, color=WHITE, size=8)
            c.fill = fl(NAVY); c.alignment = al("center", wrap=True); c.border = bd()
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A{ROW_H}:{get_column_letter(len(colunas))}{ROW_H}"
        for i, (_, row) in enumerate(df_tab.iterrows()):
            er = ROW_H + 1 + i
            arq = str(row.get("ARQUIVADO DEFINITIVAMENTE", ""))
            bg = ("FFE8E8" if i % 2 == 0 else "FFD5D5") if "SIM" in arq else \
                 ("E8F5E9" if i % 2 == 0 else "D5EED8") if "NÃO" in arq else \
                 ("FFF8E8" if i % 2 == 0 else "FFF0D0")
            for ci, col in enumerate(colunas, 1):
                c = ws.cell(row=er, column=ci); c.value = row[col]
                c.border = bd(); c.fill = fl(bg); c.font = ft(size=8)
                c.alignment = al("center" if ci <= 3 else "left")
                if col == "ARQUIVADO DEFINITIVAMENTE":
                    cor_f = RED if "SIM" in arq else (GREEN if "NÃO" in arq else AMBER)
                    c.font = Font(name="Arial", bold=True, size=8, color=cor_f)
                    c.alignment = al("center")
                if col == "Status":
                    cor_s = {"ARQUIVADO": RED, "EM ANDAMENTO": GREEN, "NÃO ENCONTRADO": AMBER,
                             "ACESSO NEGADO": AMBER, "SEM CNJ PADRÃO": BLUE}.get(str(row[col]), AMBER)
                    c.font = Font(name="Arial", bold=True, size=8, color=cor_s)
            ws.row_dimensions[er].height = 15

    ws1 = wb.active; ws1.title = "Todos os Resultados"
    escrever_aba(ws1, "SEBRAE CONTENCIOSO — RESULTADO COMPLETO CONSULTA DATAJUD/CNJ", df_res, NAVY)

    ws2 = wb.create_sheet("✅ Arquivados Definitivamente")
    if len(arquivados):
        escrever_aba(ws2, "PROCESSOS ARQUIVADOS DEFINITIVAMENTE", arquivados, RED)
    else:
        ws2["A1"] = "Nenhum processo arquivado definitivamente identificado."
        ws2["A1"].font = ft(bold=True)

    ws3 = wb.create_sheet("❌ Em Andamento")
    escrever_aba(ws3, "PROCESSOS EM ANDAMENTO CONFIRMADOS", em_andamento, GREEN)

    ws4 = wb.create_sheet("❓ Verificar Manualmente")
    if len(verificar):
        escrever_aba(ws4, "STATUS INDEFINIDO — VERIFICAR MANUALMENTE NOS PORTAIS DOS TRIBUNAIS", verificar, AMBER)

    ws5 = wb.create_sheet("📊 Resumo")
    ws5.sheet_view.showGridLines = False
    ws5.merge_cells("A1:C1")
    c = ws5["A1"]; c.value = "RESUMO — CONSULTA DATAJUD CNJ"
    c.font = Font(name="Arial", bold=True, color=WHITE, size=13)
    c.fill = fl(NAVY); c.alignment = al("center"); ws5.row_dimensions[1].height = 28
    resumo = [
        ("Total consultados", len(df_res), NAVY),
        ("✅ Arquivados definitivamente", len(arquivados), RED),
        ("❌ Em andamento (confirmados)", len(em_andamento), GREEN),
        ("❓ Verificar manualmente", len(verificar), AMBER),
        ("Data/hora da consulta", agora, BLUE),
        ("Fonte", "API DataJud/CNJ", BLUE),
    ]
    for ri, (desc, val, cor) in enumerate(resumo):
        er = 3 + ri; ws5.row_dimensions[er].height = 24
        ws5.merge_cells(f"A{er}:B{er}")
        c1 = ws5.cell(row=er, column=1); c1.value = desc
        c1.font = ft(bold=True, size=10); c1.fill = fl(GRAY_L); c1.border = bd(); c1.alignment = al("left")
        c2 = ws5.cell(row=er, column=3); c2.value = val
        c2.font = Font(name="Arial", bold=True, size=12, color=cor)
        c2.fill = fl(GRAY_L); c2.border = bd(); c2.alignment = al("center")
    for ci, w in enumerate([32, 8, 18], 1):
        ws5.column_dimensions[get_column_letter(ci)].width = w

    ws5.merge_cells("A10:C10")
    c = ws5["A10"]; c.value = "PORTAIS PARA CONSULTA MANUAL (processos sem CNJ padrão)"
    c.font = Font(name="Arial", bold=True, color=WHITE, size=10)
    c.fill = fl(BLUE); c.alignment = al("center"); ws5.row_dimensions[10].height = 18
    portais = [
        ("AD / DF", "TJDFT / TRF1", "https://www.tjdft.jus.br  |  https://processual.trf1.jus.br"),
        ("SP", "TJSP / TRF3", "https://esaj.tjsp.jus.br  |  https://www.trf3.jus.br"),
        ("RJ", "TJRJ / TRF2", "https://www3.tjrj.jus.br  |  https://eproc.trf2.jus.br"),
        ("MG", "TJMG / TRF1", "https://www.tjmg.jus.br  |  https://processual.trf1.jus.br"),
        ("PR / SC / RS", "TRF4", "https://eproc.trf4.jus.br"),
        ("BA / AM / GO / ES / MS", "TRF1", "https://processual.trf1.jus.br"),
    ]
    for ri, (uf, trib, url) in enumerate(portais):
        er = 11 + ri
        for ci, val in enumerate([uf, trib, url], 1):
            c = ws5.cell(row=er, column=ci); c.value = val
            c.font = ft(size=9); c.border = bd()
            c.fill = fl(GRAY_L if ri % 2 == 0 else WHITE); c.alignment = al("left")
        ws5.row_dimensions[er].height = 15

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
